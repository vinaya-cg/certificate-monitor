import boto3
import openpyxl
from datetime import datetime
import uuid

# Initialize DynamoDB client
dynamodb = boto3.resource('dynamodb', region_name='eu-west-1')
table = dynamodb.Table('cert-management-dev-certificates')

# Load the Excel file
wb = openpyxl.load_workbook('dummy_certificates_100.xlsx')
ws = wb.active

print("Clearing existing certificates from DynamoDB...")

# Scan and delete all existing items
try:
    scan = table.scan()
    with table.batch_writer() as batch:
        for item in scan['Items']:
            batch.delete_item(Key={'CertificateID': item['CertificateID']})
    print(f"Deleted {len(scan['Items'])} existing certificates")
except Exception as e:
    print(f"Error clearing table: {e}")

print("\nUploading 100 new dummy certificates...")

# Upload new certificates from Excel
uploaded_count = 0
with table.batch_writer() as batch:
    for row_num in range(2, ws.max_row + 1):  # Skip header row
        cert_name = ws.cell(row=row_num, column=1).value
        environment = ws.cell(row=row_num, column=2).value
        application = ws.cell(row=row_num, column=3).value
        expiry_date = ws.cell(row=row_num, column=4).value
        owner = ws.cell(row=row_num, column=5).value
        status = ws.cell(row=row_num, column=6).value
        days_left = ws.cell(row=row_num, column=7).value
        
        # Create certificate item
        cert_id = str(uuid.uuid4())
        
        item = {
            'CertificateID': cert_id,
            'CertificateName': cert_name,
            'Environment': environment,
            'ApplicationName': application,
            'ExpiryDate': expiry_date,
            'Owner': owner,
            'Status': status,
            'DaysUntilExpiry': int(days_left),
            'CreatedAt': datetime.now().isoformat(),
            'UpdatedAt': datetime.now().isoformat(),
            'CertificateType': cert_name.split('-')[2] if '-' in cert_name else 'SSL',
            'Issuer': 'PostNL CA',
            'Subject': f'CN={cert_name}',
            'SerialNumber': f'SN-{row_num-1:05d}'
        }
        
        batch.put_item(Item=item)
        uploaded_count += 1
        
        if uploaded_count % 10 == 0:
            print(f"Uploaded {uploaded_count} certificates...")

print(f"\n✅ Successfully uploaded {uploaded_count} dummy certificates to DynamoDB!")
print("\nCertificate distribution:")
print("- Expired: ~10")
print("- Due for Renewal: ~15")
print("- Renewal in Progress: ~5")
print("- Active: ~70")
