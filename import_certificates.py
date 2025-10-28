#!/usr/bin/env python3
"""
Certificate Import Script for PostNL Certificate Management System

This script reads certificates from an Excel file and imports them into DynamoDB.
Supports dry-run mode for previewing data before import.

Usage:
    python import_certificates.py --file "path/to/excel.xlsx" --dry-run
    python import_certificates.py --file "path/to/excel.xlsx" --table cert-management-dev-certificates
"""

import argparse
import boto3
import openpyxl
import uuid
from datetime import datetime
from dateutil import parser
import sys
import json
from botocore.exceptions import ClientError, NoCredentialsError


class CertificateImporter:
    def __init__(self, table_name, region='eu-west-1'):
        self.table_name = table_name
        self.region = region
        self.dynamodb = None
        self.table = None
        
    def connect_to_dynamodb(self):
        """Initialize DynamoDB connection"""
        try:
            self.dynamodb = boto3.resource('dynamodb', region_name=self.region)
            self.table = self.dynamodb.Table(self.table_name)
            # Test connection
            self.table.table_status
            print(f"✅ Connected to DynamoDB table: {self.table_name} in {self.region}")
            return True
        except NoCredentialsError:
            print("❌ AWS credentials not found. Please configure AWS CLI or set environment variables.")
            return False
        except ClientError as e:
            print(f"❌ Error connecting to DynamoDB: {e}")
            return False
        except Exception as e:
            print(f"❌ Unexpected error: {e}")
            return False
    
    def parse_date(self, date_value):
        """Parse various date formats and return ISO string"""
        if not date_value:
            return None
            
        try:
            # Handle Excel datetime objects
            if hasattr(date_value, 'strftime'):
                return date_value.strftime('%Y-%m-%d')
            
            # Handle string dates
            if isinstance(date_value, str):
                # Try parsing with dateutil
                parsed_date = parser.parse(date_value)
                return parsed_date.strftime('%Y-%m-%d')
                
        except Exception as e:
            print(f"⚠️  Warning: Could not parse date '{date_value}': {e}")
            return str(date_value) if date_value else None
    
    def read_excel_file(self, file_path):
        """Read certificates from Excel file"""
        try:
            print(f"📖 Reading Excel file: {file_path}")
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value.strip())
            
            print(f"📋 Found {len(headers)} columns: {headers}")
            print(f"📊 Total rows in Excel: {worksheet.max_row}")
            
            certificates = []
            skipped_rows = 0
            
            # Process data rows (skip header)
            for row_num in range(2, worksheet.max_row + 1):
                row_data = {}
                has_data = False
                
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    
                    if value is not None:
                        has_data = True
                        if isinstance(value, str):
                            value = value.strip()
                        row_data[header] = value
                
                # Skip completely empty rows
                if not has_data:
                    skipped_rows += 1
                    continue
                
                # Generate unique CertificateID
                certificate_id = f"cert-{str(uuid.uuid4())[:8]}"
                
                # Map Excel columns to DynamoDB attributes
                certificate = {
                    'CertificateID': certificate_id,
                    'SerialNumber': str(row_data.get('Serial Number', '')),
                    'CertificateName': str(row_data.get('Certificate  Name', row_data.get('Certificate Name', ''))),
                    'ExpiryDate': self.parse_date(row_data.get('Expiry Date')),
                    'AccountNumber': str(row_data.get('AccountNumber', '')),
                    'Application': str(row_data.get('Application', '')),
                    'Environment': str(row_data.get('ENVIRONMENT', '')),
                    'Type': str(row_data.get('Type', '')),
                    'Status': str(row_data.get('Status', 'Unknown')),
                    'OwnerEmail': str(row_data.get('OwnerEmail', '')),
                    'SupportEmail': str(row_data.get('Support team Email', '')),
                    'LastUpdatedOn': datetime.utcnow().isoformat() + 'Z'
                }
                
                # Add optional fields if present
                if row_data.get('IncidentNumber'):
                    certificate['IncidentNumber'] = str(row_data['IncidentNumber'])
                if row_data.get('RenewedBy'):
                    certificate['RenewedBy'] = str(row_data['RenewedBy'])
                if row_data.get('RenewalLog'):
                    certificate['RenewalLog'] = str(row_data['RenewalLog'])
                if row_data.get('UploadS3Key'):
                    certificate['UploadS3Key'] = str(row_data['UploadS3Key'])
                
                certificates.append(certificate)
            
            print(f"✅ Processed {len(certificates)} certificates ({skipped_rows} empty rows skipped)")
            return certificates
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return []
    
    def preview_certificates(self, certificates, limit=5):
        """Preview certificates before import"""
        print(f"\n📋 PREVIEW: First {min(limit, len(certificates))} certificates:")
        print("=" * 80)
        
        for i, cert in enumerate(certificates[:limit]):
            print(f"\n🔖 Certificate {i+1}:")
            print(f"   ID: {cert['CertificateID']}")
            print(f"   Name: {cert['CertificateName']}")
            print(f"   Environment: {cert['Environment']}")
            print(f"   Application: {cert['Application']}")
            print(f"   Expiry: {cert['ExpiryDate']}")
            print(f"   Status: {cert['Status']}")
            print(f"   Owner: {cert['OwnerEmail']}")
        
        if len(certificates) > limit:
            print(f"\n... and {len(certificates) - limit} more certificates")
        
        print("=" * 80)
    
    def import_certificates(self, certificates, batch_size=25):
        """Import certificates to DynamoDB in batches"""
        if not certificates:
            print("❌ No certificates to import")
            return False
        
        print(f"\n🚀 Starting import of {len(certificates)} certificates...")
        success_count = 0
        error_count = 0
        
        # Process in batches (DynamoDB batch_writer limit is 25)
        for i in range(0, len(certificates), batch_size):
            batch = certificates[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (len(certificates) + batch_size - 1) // batch_size
            
            print(f"📦 Processing batch {batch_num}/{total_batches} ({len(batch)} items)...")
            
            try:
                with self.table.batch_writer() as batch_writer:
                    for cert in batch:
                        # Remove empty string values to avoid DynamoDB issues
                        clean_cert = {k: v for k, v in cert.items() if v != ''}
                        batch_writer.put_item(Item=clean_cert)
                        success_count += 1
                
                print(f"✅ Batch {batch_num} completed successfully")
                
            except Exception as e:
                print(f"❌ Error in batch {batch_num}: {e}")
                error_count += len(batch)
        
        print(f"\n📊 Import Summary:")
        print(f"   ✅ Successfully imported: {success_count}")
        print(f"   ❌ Failed: {error_count}")
        print(f"   📈 Success rate: {(success_count/(success_count+error_count)*100):.1f}%")
        
        return success_count > 0


def main():
    parser = argparse.ArgumentParser(description='Import certificates from Excel to DynamoDB')
    parser.add_argument('--file', required=True, help='Path to Excel file')
    parser.add_argument('--table', default='cert-management-dev-certificates', help='DynamoDB table name')
    parser.add_argument('--region', default='eu-west-1', help='AWS region')
    parser.add_argument('--dry-run', action='store_true', help='Preview data without importing')
    parser.add_argument('--preview', type=int, default=5, help='Number of certificates to preview')
    
    args = parser.parse_args()
    
    print("🔧 PostNL Certificate Import Tool")
    print("=" * 50)
    
    # Initialize importer
    importer = CertificateImporter(args.table, args.region)
    
    # Read certificates from Excel
    certificates = importer.read_excel_file(args.file)
    if not certificates:
        print("❌ No certificates found or error reading file")
        sys.exit(1)
    
    # Always show preview
    importer.preview_certificates(certificates, args.preview)
    
    if args.dry_run:
        print(f"\n🔍 DRY RUN MODE - No data will be imported")
        print(f"📋 Found {len(certificates)} certificates ready for import")
        print(f"🎯 Target table: {args.table} in {args.region}")
        print(f"\nTo actually import, run without --dry-run flag")
        return
    
    # Connect to DynamoDB
    if not importer.connect_to_dynamodb():
        sys.exit(1)
    
    # Confirm import
    response = input(f"\n⚠️  Import {len(certificates)} certificates to {args.table}? (y/N): ")
    if response.lower() != 'y':
        print("❌ Import cancelled")
        return
    
    # Import certificates
    success = importer.import_certificates(certificates)
    
    if success:
        print(f"\n🎉 Import completed! Certificates are now available in your dashboard:")
        print(f"   🌐 Dashboard: http://cert-management-dev-dashboard-a3px89bh.s3-website-eu-west-1.amazonaws.com")
        print(f"\n✅ You can now:")
        print(f"   - View all certificates in the dashboard")
        print(f"   - Update certificate status")
        print(f"   - Upload renewal documents")
        print(f"   - Receive automated expiry notifications")
    else:
        print("❌ Import failed")
        sys.exit(1)


if __name__ == "__main__":
    main()