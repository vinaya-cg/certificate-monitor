import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Certificates"

# Define headers
headers = [
    "Certificate Name",
    "Environment",
    "Application Name",
    "Expiry Date",
    "Owner",
    "Status",
    "Days Until Expiry"
]

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Define data for variety
environments = ["Production", "Development", "Staging", "UAT", "QA"]
applications = [
    "API Gateway", "Web Portal", "Mobile App", "Database Service", "Payment Gateway",
    "Authentication Service", "Email Service", "File Storage", "Analytics Platform",
    "Reporting Service", "CRM System", "ERP System", "HR Portal", "Customer Portal",
    "Admin Dashboard", "Monitoring Service", "Backup Service", "Load Balancer",
    "CDN Service", "VPN Gateway"
]
owners = [
    "john.doe@postnl.com", "jane.smith@postnl.com", "mike.johnson@postnl.com",
    "sarah.williams@postnl.com", "david.brown@postnl.com", "emily.davis@postnl.com",
    "robert.miller@postnl.com", "lisa.wilson@postnl.com", "james.moore@postnl.com",
    "maria.taylor@postnl.com"
]

cert_types = ["SSL", "TLS", "API", "Code Signing", "Email", "VPN"]

# Generate 100 dummy certificates
base_date = datetime.now()
for i in range(1, 101):
    # Certificate name
    cert_type = random.choice(cert_types)
    app = random.choice(applications)
    env_short = random.choice(environments)[:3].upper()
    cert_name = f"{app.replace(' ', '')}-{env_short}-{cert_type}-{i:03d}"
    
    # Environment
    environment = random.choice(environments)
    
    # Application
    application = app
    
    # Expiry date - varied distribution
    if i <= 10:  # 10 expired
        days_offset = random.randint(-180, -1)
    elif i <= 25:  # 15 due for renewal (1-30 days)
        days_offset = random.randint(1, 30)
    elif i <= 30:  # 5 renewal in progress (31-60 days)
        days_offset = random.randint(31, 60)
    else:  # 70 active (60+ days)
        days_offset = random.randint(61, 730)
    
    expiry_date = base_date + timedelta(days=days_offset)
    
    # Owner
    owner = random.choice(owners)
    
    # Status - calculate based on days
    if days_offset < 0:
        status = "Expired"
    elif days_offset <= 30:
        status = "Due for Renewal"
    elif days_offset <= 60:
        status = "Renewal in Progress" if random.random() > 0.5 else "Active"
    else:
        status = "Active"
    
    # Days until expiry
    days_left = days_offset
    
    # Write row
    row = i + 1
    ws.cell(row=row, column=1).value = cert_name
    ws.cell(row=row, column=2).value = environment
    ws.cell(row=row, column=3).value = application
    ws.cell(row=row, column=4).value = expiry_date.strftime("%Y-%m-%d")
    ws.cell(row=row, column=5).value = owner
    ws.cell(row=row, column=6).value = status
    ws.cell(row=row, column=7).value = days_left
    
    # Apply conditional formatting for status
    status_cell = ws.cell(row=row, column=6)
    if status == "Expired":
        status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_cell.font = Font(color="9C0006")
    elif status == "Due for Renewal":
        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        status_cell.font = Font(color="9C5700")
    elif status == "Renewal in Progress":
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        status_cell.font = Font(color="006100")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
filename = "dummy_certificates_100.xlsx"
wb.save(filename)
print(f"Created {filename} with 100 dummy certificates")
print(f"Distribution: ~10 Expired, ~15 Due for Renewal, ~5 Renewal in Progress, ~70 Active")
