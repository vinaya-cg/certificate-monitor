#!/usr/bin/env python3
"""
Quick test to verify Python environment and AWS connectivity
"""

import boto3
from decimal import Decimal

def test_dynamodb_connection():
    """Test connection to DynamoDB and read certificate data"""
    print("🔍 Testing AWS DynamoDB connection...")
    
    try:
        # Create DynamoDB resource
        dynamodb = boto3.resource('dynamodb', region_name='eu-west-1')
        table = dynamodb.Table('cert-management-dev-certificates')
        
        print("✅ Connected to DynamoDB!")
        print(f"✅ Table: {table.table_name}")
        
        # Scan for a few certificates
        response = table.scan(Limit=3)
        count = response.get('Count', 0)
        
        print(f"✅ Retrieved {count} sample certificates")
        
        if count > 0:
            print("\n📋 Sample Certificate Data:")
            for i, cert in enumerate(response['Items'][:3], 1):
                print(f"\n  Certificate {i}:")
                print(f"    ID: {cert.get('CertificateID', 'N/A')}")
                print(f"    Common Name: {cert.get('CommonName', 'N/A')}")
                print(f"    Expiry Date: {cert.get('ExpiryDate', 'N/A')}")
                print(f"    Status: {cert.get('Status', 'N/A')}")
                print(f"    Environment: {cert.get('Environment', 'N/A')}")
        
        # Get total count
        print("\n🔢 Getting total certificate count...")
        scan_response = table.scan(Select='COUNT')
        total = scan_response.get('Count', 0)
        print(f"✅ Total certificates in DynamoDB: {total}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error connecting to DynamoDB: {e}")
        return False

def test_excel_capability():
    """Test openpyxl is working"""
    print("\n📊 Testing Excel file handling...")
    
    try:
        from openpyxl import Workbook
        
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test Certificate"
        ws['B1'] = "2026-12-31"
        
        print("✅ openpyxl can create workbooks")
        print("✅ openpyxl can write data")
        
        return True
        
    except Exception as e:
        print(f"❌ Error testing openpyxl: {e}")
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 TESTING PYTHON ENVIRONMENT WITH REAL AWS DATA")
    print("=" * 60)
    print()
    
    # Test DynamoDB
    db_test = test_dynamodb_connection()
    
    # Test Excel
    excel_test = test_excel_capability()
    
    print("\n" + "=" * 60)
    if db_test and excel_test:
        print("✅ ALL TESTS PASSED - Environment is fully functional!")
    else:
        print("❌ Some tests failed - check errors above")
    print("=" * 60)
